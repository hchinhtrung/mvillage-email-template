import openpyxl
import re
from pathlib import Path

def detect_hr_pattern_in_excel(file_path):
    """
    Detect cells containing pattern HRxxxxxx (HR followed by digits)
    in an Excel file.
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        List of dictionaries containing sheet name, cell address, and cell value
    """
    # Pattern: HR followed by one or more digits
    pattern = re.compile(r'HR\d+', re.IGNORECASE)
    
    results = []
    
    # Load the workbook
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # Iterate through all sheets
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        # Iterate through all cells in the sheet
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value:
                    # Convert cell value to string and search for pattern
                    cell_str = str(cell.value)
                    if pattern.search(cell_str):
                        results.append({
                            'file': Path(file_path).name,
                            'sheet': sheet_name,
                            'cell': cell.coordinate,
                            'value': cell.value
                        })
    
    return results


def main():
    # List of uploaded files
    files = [
        '/Users/hchinhtrung/Documents/GitHub/mvillage-email-template/15.python/Best of Vietnam for Teenager.xlsx',
        '/Users/hchinhtrung/Documents/GitHub/mvillage-email-template/15.python/Vietnam Explorer for teenager.xlsx',
        '/Users/hchinhtrung/Documents/GitHub/mvillage-email-template/15.python/Vietnam Family holiday.xlsx'
    ]
    
    all_results = []
    
    # Process each file
    for file_path in files:
        print(f"\n{'='*60}")
        print(f"Processing: {Path(file_path).name}")
        print(f"{'='*60}")
        
        try:
            results = detect_hr_pattern_in_excel(file_path)
            
            if results:
                print(f"Found {len(results)} cell(s) with HR pattern:\n")
                for result in results:
                    print(f"  Sheet: {result['sheet']}")
                    print(f"  Cell:  {result['cell']}")
                    print(f"  Value: {result['value']}")
                    print()
                all_results.extend(results)
            else:
                print("No cells with HR pattern found.")
                
        except Exception as e:
            print(f"Error processing file: {e}")
    
    # Summary
    print(f"\n{'='*60}")
    print(f"SUMMARY: Found {len(all_results)} total cell(s) with HR pattern")
    print(f"{'='*60}")
    
    return all_results


if __name__ == "__main__":
    results = main()
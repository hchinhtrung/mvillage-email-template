# -*- coding: utf-8 -*-
"""Hotel-list input (Google Sheet master / CSV / XLSX) + auto-sharding.

Input resolution:
  - "gsheet:<id>" or a docs.google.com URL  -> Google Sheet via gviz CSV export
  - "*.xlsx"                                 -> Excel (sheet with hyperlink targets)
  - anything else                            -> CSV (searched under 31.crawl-tool if not found)

Only the first three columns are used (hotel_name, hotel_url, room_type); the original URL
host is preserved so returned room names match the language of room_type in the sheet.
"""
import glob
import os
import re
from urllib.parse import quote


def _norm_cols(df):
    req = ["hotel_name", "hotel_url", "room_type"]
    if not all(c in df.columns for c in req):
        df.columns = req + list(df.columns[3:])
    df = df[df["hotel_url"].notna() & (df["hotel_url"].astype(str).str.strip() != "")]
    df = df[req]
    bad = df[~df["hotel_url"].astype(str).str.strip().str.startswith(("http://", "https://"))]
    if len(bad):
        raise ValueError(
            f"column 2 (hotel_url) is not a URL in {len(bad)}/{len(df)} rows "
            f"(e.g. {str(bad.iloc[0]['hotel_url'])[:60]!r}). Input needs the first three columns "
            f"to be (hotel_name, hotel_url, room_type) — TEMP_*/FINAL_* files are crawl OUTPUT, "
            f"not input.")
    return df


def _resolve_local(path):
    if os.path.exists(path):
        return path
    here = os.path.dirname(os.path.abspath(__file__))
    root = os.path.dirname(here)  # 31.crawl-tool
    cands = glob.glob(os.path.join(root, "**", os.path.basename(path)), recursive=True)
    return cands[0] if cands else path


def is_gsheet(spec):
    spec = str(spec)
    return spec.startswith("gsheet:") or "docs.google.com" in spec


def _parse_gsheet(spec):
    """Return (spreadsheet_id, gid) from a full Sheets URL, a 'gsheet:<id>[:<gid>]', or a bare id."""
    sid, gid = spec, ""
    if spec.startswith("gsheet:"):
        rest = spec.split(":", 1)[1]
        sid, _, gid = rest.partition(":")
    elif "docs.google.com" in spec:
        m = re.search(r"/d/([A-Za-z0-9_-]+)", spec)
        if m:
            sid = m.group(1)
        g = re.search(r"[?#&]gid=(\d+)", spec)      # the specific tab (worksheet id)
        if g:
            gid = g.group(1)
    return sid, gid


def _gsheet_url(spec, sheet_name="", gid=""):
    sid, url_gid = _parse_gsheet(spec)
    gid = gid or url_gid
    url = f"https://docs.google.com/spreadsheets/d/{sid}/gviz/tq?tqx=out:csv"
    if gid:
        url += f"&gid={gid}"          # gid takes precedence — it targets the exact tab
    elif sheet_name:
        url += f"&sheet={quote(sheet_name)}"
    return url


def read_hotels(input_spec, sheet_name="", gid=""):
    """Return a list of (hotel_name, hotel_url, room_type) tuples."""
    import pandas as pd

    spec = str(input_spec)
    if is_gsheet(spec):
        df = _norm_cols(pd.read_csv(_gsheet_url(spec, sheet_name, gid)))
    elif spec.endswith(".xlsx"):
        df = _read_xlsx(_resolve_local(spec), sheet_name or "Hotel Link")
    else:
        df = _norm_cols(pd.read_csv(_resolve_local(spec), encoding="utf-8-sig"))
    return [(r["hotel_name"], r["hotel_url"], r["room_type"]) for _, r in df.iterrows()]


def _read_xlsx(path, sheet_name):
    import pandas as pd
    from openpyxl import load_workbook

    wb = load_workbook(filename=path, data_only=True)
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=2, max_col=3):
        hcell = row[0]
        rcell = row[2] if len(row) > 2 else None
        url = hcell.hyperlink.target if hcell.hyperlink else (hcell.value or "")
        rows.append({"hotel_name": hcell.value or "", "hotel_url": url,
                     "room_type": (rcell.value if rcell else "")})
    df = pd.DataFrame(rows)
    df = df[df["hotel_url"].notna() & df["hotel_url"].astype(str).str.startswith("http")]
    return df[["hotel_name", "hotel_url", "room_type"]]


def parse_shard(spec):
    """'2/5' -> (2, 5). Returns None if spec is empty/invalid."""
    if not spec:
        return None
    try:
        n, m = str(spec).split("/")
        n, m = int(n), int(m)
        if 1 <= n <= m:
            return (n, m)
    except Exception:
        pass
    raise ValueError(f"--shard must be 'N/M' with 1<=N<=M, got {spec!r}")


def apply_shard(items, shard):
    """Contiguous split into M near-equal chunks; return chunk N (1-based). Contiguous (not
    modulo) so a shard's rows stay grouped, which keeps resume/checkpoint local."""
    if not shard:
        return items
    n, m = shard
    total = len(items)
    base, rem = divmod(total, m)
    # chunk i (0-based) gets base (+1 for the first `rem` chunks)
    start = 0
    for i in range(n - 1):
        start += base + (1 if i < rem else 0)
    size = base + (1 if (n - 1) < rem else 0)
    return items[start:start + size]

from openpyxl import load_workbook, Workbook
import os
import math
import re
from datetime import datetime

# =====================
# CONFIG
# =====================
# Get the directory where the script is located
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

INPUT_FILE = os.path.join(SCRIPT_DIR, "raw.xlsx")
ROWS_PER_FILE = 5
LOCATION_COL = 1   # cột A (chỉ dùng để group)
KEEP_COLS = [2, 3] # chỉ giữ cột B, C
OUTPUT_ROOT = os.path.join(SCRIPT_DIR, "export")

# =====================
# UTILS
# =====================
def sanitize(text):
    return re.sub(r'[\\/:*?"<>| ]+', '_', str(text).strip())

# =====================
# MAIN
# =====================
def export_by_location():
    wb = load_workbook(INPUT_FILE)
    ws = wb.active

    # Header: chỉ lấy B & C
    header = [ws.cell(row=1, column=c).value for c in KEEP_COLS]

    rows_by_location = {}

    for row in ws.iter_rows(min_row=2):
        location = row[LOCATION_COL - 1].value
        if not location:
            continue
        rows_by_location.setdefault(location, []).append(row)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    base_output = os.path.join(OUTPUT_ROOT, f"Export_{timestamp}")
    os.makedirs(base_output, exist_ok=True)

    total_files = 0

    for location, rows in rows_by_location.items():
        loc_name = sanitize(location)
        loc_dir = os.path.join(base_output, loc_name)
        os.makedirs(loc_dir, exist_ok=True)

        chunks = math.ceil(len(rows) / ROWS_PER_FILE)

        for i in range(chunks):
            chunk = rows[i * ROWS_PER_FILE:(i + 1) * ROWS_PER_FILE]

            new_wb = Workbook()
            new_ws = new_wb.active
            new_ws.title = "Hotel Link"

            # Write header (B, C)
            new_ws.append(header)

            for row in chunk:
                # write values B, C
                values = [row[c - 1].value for c in KEEP_COLS]
                new_ws.append(values)

                # copy hyperlink (chỉ cột B)
                hotel_cell = row[KEEP_COLS[0] - 1]
                if hotel_cell.hyperlink:
                    new_ws.cell(
                        row=new_ws.max_row,
                        column=1
                    ).hyperlink = hotel_cell.hyperlink.target

            output_file = os.path.join(
                loc_dir,
                f"{loc_name}-{i + 1}.xlsx"
            )
            new_wb.save(output_file)
            total_files += 1

    print(f"✅ DONE — Total files: {total_files}")
    print(f"📁 Output folder: {os.path.abspath(base_output)}")

# =====================
# RUN
# =====================
if __name__ == "__main__":
    export_by_location()
